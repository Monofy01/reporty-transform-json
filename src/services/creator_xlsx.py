import io
import json

from openpyxl.workbook import Workbook

from src.models.excel import Excel
from src.models.sheet import Sheet


class CreatorXlsx:
    def __init__(self, body):
        self.workbook = Workbook()
        self.excel_raw = body
        self.log_output = []
        self.excel = Excel(
            filename=json.loads(body['excel'])['filename'],
            webhook=json.loads(body['excel'])['webhook'],
            sheets=[Sheet.from_dict(item) for item in json.loads(body['excel'])['sheets']]
        )


    def new_xlsx(self):
        for sheet_file in self.excel.sheets:
            sheet = self.workbook.create_sheet(title=sheet_file.name)

            for column_index, column_name in enumerate(sheet_file.columns, start=1):
                sheet.cell(row=1, column=column_index, value=column_name[0])

            for row_index, row_data in enumerate(sheet_file.data, start=2):
                for column_index, column_name in enumerate(sheet_file.columns, start=1):
                    cell_value = row_data.get(column_name[0])
                    if type(cell_value) is list or type(cell_value) is bool:
                        sheet.cell(row=row_index, column=column_index, value=str(cell_value))
                    else:
                        sheet.cell(row=row_index, column=column_index, value=cell_value)
            if len(sheet_file.data_invalid) > 0:
                self.log_output.append([dict(sheet_file.columns), *sheet_file.data_invalid])

        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)

        binary_data = buffer.getvalue()
        buffer.close()
        return binary_data

